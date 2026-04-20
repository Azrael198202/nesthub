from __future__ import annotations

import json
import os
import mimetypes
import re
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

from nethub_runtime.core.config.settings import SEMANTIC_POLICY_PATH
from nethub_runtime.core.memory.semantic_policy_store import SemanticPolicyStore
from nethub_runtime.core.schemas.context_schema import CoreContextSchema
from nethub_runtime.core.schemas.task_schema import SubTask, TaskSchema
from nethub_runtime.core.schemas.workflow_schema import WorkflowSchema, WorkflowStepSchema
from nethub_runtime.core.services.execution_handler_registry import (
    ExecutionHandlerPluginManifest,
    ExecutionHandlerPluginStepSpec,
)
from nethub_runtime.core.services.execution_step_handlers import extract_image_text_with_ocr
from nethub_runtime.core.utils.id_generator import generate_id

_DOCUMENT_SUFFIXES = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt", ".md"}
_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"}
_SUPPORTED_SUFFIXES = _DOCUMENT_SUFFIXES | _IMAGE_SUFFIXES
_POLICY_STORE = SemanticPolicyStore(policy_path=SEMANTIC_POLICY_PATH)


def _document_runtime_policy() -> dict[str, Any]:
    try:
        return _POLICY_STORE.load_runtime_policy().get("document_runtime") or {}
    except Exception:
        return {}


def _policy_markers(key: str) -> tuple[str, ...]:
    value = _document_runtime_policy().get(key) or []
    return tuple(str(item) for item in value if str(item).strip())


def _utc_now() -> datetime:
    return datetime.now(UTC)


def _document_wait_timeout_seconds() -> int:
    raw = str(os.getenv("NESTHUB_DOCUMENT_WAIT_TIMEOUT_SECONDS", "300") or "300").strip()
    try:
        parsed = int(raw)
    except ValueError:
        parsed = 300
    return max(120, min(parsed, 300))


def _is_bridge_line_context(context: CoreContextSchema) -> bool:
    return str((context.metadata or {}).get("source_im") or "").strip().lower() == "line"


def _build_standby_message(wait_state: str, context: CoreContextSchema) -> str:
    if _is_bridge_line_context(context):
        return ""
    if wait_state == "awaiting_document":
        return "已记录文档处理需求，等待文档或图片上传后自动继续。"
    return "已收到文档或图片，等待你的处理要求后继续。"


def _parse_iso_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        normalized = text.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def _pending_request_is_active(payload: dict[str, Any] | None) -> bool:
    if not isinstance(payload, dict):
        return False
    request_text = str(payload.get("request_text") or "").strip()
    if not request_text:
        return False
    expires_at = _parse_iso_datetime(payload.get("expires_at"))
    if expires_at is None:
        return False
    return expires_at >= _utc_now()


def _get_active_pending_request(state: dict[str, Any]) -> dict[str, Any] | None:
    payload = state.get("pending_document_request")
    if not isinstance(payload, dict):
        return None
    return payload if _pending_request_is_active(payload) else None


def _looks_like_explicit_document_request(text: str, *, has_context: bool) -> bool:
    lowered = text.lower()
    if not lowered.strip():
        return False
    summary_markers = _policy_markers("summary_markers")
    translate_markers = _policy_markers("translate_markers")
    analyze_markers = _policy_markers("analyze_markers")
    action_requested = (
        any(marker in lowered for marker in summary_markers)
        or any(marker in lowered for marker in translate_markers)
        or any(marker in lowered for marker in analyze_markers)
    )
    if not action_requested:
        return False
    document_reference_markers = _policy_markers("document_reference_markers")
    followup_reference_markers = _policy_markers("followup_reference_markers")
    has_document_reference = (
        any(marker in lowered for marker in document_reference_markers)
        or any(marker in text for marker in followup_reference_markers)
        or any(_looks_like_document_name(token) for token in re.findall(r"[\w.-]+", lowered))
    )
    return has_document_reference or has_context


def _resolve_analysis_targets(text: str, context: CoreContextSchema) -> list[dict[str, Any]]:
    attachments = list((context.session_state or {}).get("analysis_attachments") or [])
    if not attachments:
        attachments = list((context.session_state or {}).get("documents") or [])
    if not attachments:
        return []
    lowered = text.lower()
    if any(marker in lowered for marker in _policy_markers("multi_doc_markers")):
        return attachments
    named_matches = []
    for item in attachments:
        file_name = str(item.get("file_name") or "")
        if file_name and file_name.lower() in lowered:
            named_matches.append(item)
    if named_matches:
        return named_matches
    return [attachments[-1]]


def _looks_like_document_name(value: str) -> bool:
    lowered = value.lower()
    return any(lowered.endswith(suffix) for suffix in _SUPPORTED_SUFFIXES)


def _detect_document_action(text: str) -> str:
    """Return the primary document action for *text*.

    Compound actions take priority over single actions:
    - summarize + translate + output-file → "summarize_translate_save"
    - summarize + translate              → "summarize_and_translate"
    Single actions are checked in priority order: translate > summarize > analyze.
    """
    lowered = text.lower()
    has_summarize = any(m in lowered for m in _policy_markers("summary_markers"))
    has_translate = any(m in lowered for m in _policy_markers("translate_markers"))
    if has_summarize and has_translate:
        if _detect_output_file_format(text) is not None:
            return "summarize_translate_save"
        return "summarize_and_translate"
    if has_translate:
        return "translate"
    if has_summarize:
        return "summarize"
    if any(marker in lowered for marker in _policy_markers("analyze_markers")):
        return "analyze"
    return "summarize"


_COMPOUND_ACTIONS = frozenset({"summarize_translate_save", "summarize_and_translate"})


def _looks_like_visual_analysis_request(text: str) -> bool:
    lowered = text.lower()
    return any(marker in lowered for marker in _policy_markers("visual_analyze_markers"))


def _is_image_target(item: dict[str, Any]) -> bool:
    input_type = str(item.get("input_type") or "").lower()
    content_type = str(item.get("content_type") or "").lower()
    file_name = str(item.get("file_name") or "").lower()
    return input_type == "image" or content_type.startswith("image/") or any(file_name.endswith(suffix) for suffix in _IMAGE_SUFFIXES)


def _image_to_data_url(path: Path, content_type: str) -> str:
    resolved_type = content_type or mimetypes.guess_type(path.name)[0] or "image/png"
    encoded = path.read_bytes().hex()
    import base64

    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{resolved_type};base64,{encoded}"


def _detect_target_language(text: str, context: CoreContextSchema) -> str:
    lowered = text.lower()
    if "中文" in text or "chinese" in lowered:
        return "Chinese"
    if "英文" in text or "english" in lowered:
        return "English"
    if "日文" in text or "日语" in text or "japanese" in lowered:
        return "Japanese"
    locale = str(context.locale or "").lower()
    if locale.startswith("zh"):
        return "Chinese"
    if locale.startswith("ja"):
        return "Japanese"
    return "English"


def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _pip_install(package: str) -> bool:
    """Run `pip install <package>` in a subprocess. Returns True on success."""
    import importlib
    import logging
    import subprocess
    import sys

    logging.getLogger(__name__).info("Installing parser package via pip: %s", package)
    try:
        result = subprocess.run(
            [sys.executable, "-m", "pip", "install", "--quiet", package],
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode == 0:
            importlib.invalidate_caches()
            return True
        logging.getLogger(__name__).warning("pip install %s failed: %s", package, result.stderr.strip())
        return False
    except Exception as exc:
        logging.getLogger(__name__).warning("pip install %s error: %s", package, exc)
        return False


def _ask_llm_for_parser_package(file_suffix: str, error: Exception) -> str | None:
    """Ask Groq / Gemini / OpenAI which pip package is needed to parse *file_suffix* files.
    Returns a pip package name string, or None if all LLM calls fail.
    Uses only cloud providers so that an M2 GPU OOM on Ollama does not block this helper.
    """
    import logging
    import os

    logger = logging.getLogger(__name__)
    prompt = (
        f"A Python program failed to parse a `{file_suffix}` file.\n"
        f"Error: {type(error).__name__}: {error}\n"
        f"What is the single pip package name to install so Python can parse `{file_suffix}` files? "
        f"Reply with ONLY the pip package name — no explanation, no punctuation."
    )
    candidates: list[tuple[str, str | None]] = []
    if os.getenv("GROQ_API_KEY"):
        candidates.append(("groq/llama-3.1-8b-instant", os.getenv("GROQ_API_KEY")))
    if os.getenv("GEMINI_API_KEY"):
        candidates.append(("gemini/gemini-2.0-flash-lite", os.getenv("GEMINI_API_KEY")))
    if os.getenv("OPENAI_API_KEY"):
        candidates.append(("openai/gpt-4o-mini", os.getenv("OPENAI_API_KEY")))

    for model, api_key in candidates:
        try:
            from litellm import completion  # type: ignore

            resp = completion(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                api_key=api_key,
                timeout=15,
                max_tokens=20,
            )
            raw = resp.choices[0].message.content.strip().split("\n")[0].strip("`\"' ")
            # Accept only a plausible single-token package name
            if raw and len(raw) < 64 and " " not in raw:
                logger.info("LLM (%s) suggested parser package '%s' for %s", model, raw, file_suffix)
                return raw
        except Exception as exc:
            logger.debug("LLM %s failed for parser resolution: %s", model, exc)
            continue
    logger.warning("No LLM could suggest a parser package for %s", file_suffix)
    return None


def _resolve_and_install_parser(file_suffix: str, error: Exception) -> bool:
    """Ask an LLM for the right package and install it. Returns True if install succeeded."""
    package = _ask_llm_for_parser_package(file_suffix, error)
    if not package:
        return False
    return _pip_install(package)


def _extract_document_text(path: Path) -> tuple[str, str]:
    import importlib

    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return _read_text_file(path), "plain_text"

    if suffix == ".pdf":
        for attempt in range(2):
            try:
                importlib.invalidate_caches()
                from pypdf import PdfReader  # type: ignore

                reader = PdfReader(str(path))
                content = "\n".join((page.extract_text() or "") for page in reader.pages)
                return content, "pypdf"
            except ModuleNotFoundError as exc:
                if attempt == 0 and _resolve_and_install_parser(suffix, exc):
                    continue
                return "", "pdf_unavailable"
            except Exception:
                return "", "pdf_unavailable"

    if suffix in {".docx", ".doc"}:
        for attempt in range(2):
            try:
                importlib.invalidate_caches()
                import docx  # type: ignore

                document = docx.Document(str(path))
                return "\n".join(paragraph.text for paragraph in document.paragraphs), "python_docx"
            except ModuleNotFoundError as exc:
                if attempt == 0 and _resolve_and_install_parser(suffix, exc):
                    continue
                return "", "docx_unavailable"
            except Exception:
                return "", "docx_unavailable"

    if suffix in {".xlsx", ".xls"}:
        for attempt in range(2):
            try:
                importlib.invalidate_caches()
                from openpyxl import load_workbook  # type: ignore

                workbook = load_workbook(str(path), data_only=True)
                lines: list[str] = []
                for sheet in workbook.worksheets:
                    lines.append(f"# Sheet: {sheet.title}")
                    for row in sheet.iter_rows(values_only=True):
                        values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                        if values:
                            lines.append(" | ".join(values))
                return "\n".join(lines), "openpyxl"
            except ModuleNotFoundError as exc:
                if attempt == 0 and _resolve_and_install_parser(suffix, exc):
                    continue
                return "", "excel_unavailable"
            except Exception:
                return "", "excel_unavailable"

    # Unknown / other file types: ask LLM what to do, attempt one dynamic install + eval
    try:
        _unknown_exc = ModuleNotFoundError(f"No parser for {suffix}")
        pkg = _ask_llm_for_parser_package(suffix, _unknown_exc)
        if pkg and _pip_install(pkg):
            importlib.invalidate_caches()
            # Best-effort: try reading as plain text after install (generic fallback)
            try:
                return _read_text_file(path), f"{pkg}_text_fallback"
            except Exception:
                pass
    except Exception:
        pass

    if suffix in _IMAGE_SUFFIXES:
        return extract_image_text_with_ocr(None, path)
    return "", "unsupported"


def _fallback_summary(text: str, file_names: list[str]) -> str:
    cleaned = " ".join(text.split())
    if not cleaned:
        return f"已接收文件 {', '.join(file_names)}，但当前环境缺少可用解析器，暂时无法提取正文。"
    excerpt = cleaned[:400]
    return f"文件 {', '.join(file_names)} 的内容摘录：{excerpt}"


def _policy_output_format_patterns() -> list[tuple[str, str]]:
    """Return [(pattern, extension), ...] loaded from the policy store.

    Expected policy shape under ``document_runtime.output_format_patterns``::

        output_format_patterns:
          - pattern: "..."
            ext: ".txt"
          - [pattern, ext]   # list form also accepted
    """
    raw = _document_runtime_policy().get("output_format_patterns") or []
    result: list[tuple[str, str]] = []
    for entry in raw:
        if isinstance(entry, (list, tuple)) and len(entry) >= 2:
            result.append((str(entry[0]), str(entry[1])))
        elif isinstance(entry, dict):
            pattern = str(entry.get("pattern") or "")
            ext = str(entry.get("ext") or "")
            if pattern and ext:
                result.append((pattern, ext))
    return result


def _detect_output_file_format(text: str) -> str | None:
    """Return a file extension if the user explicitly wants output saved as a file, else None.

    All keyword matching is driven by the policy store — see
    ``document_runtime.output_format_patterns``, ``output_delivery_words``, and
    ``output_trigger_words``.
    """
    lowered = text.lower()
    for pattern, ext in _policy_output_format_patterns():
        if pattern in lowered:
            return ext
    delivery_words = _policy_markers("output_delivery_words")
    trigger_words = _policy_markers("output_trigger_words")
    if delivery_words and trigger_words:
        if any(w in lowered for w in delivery_words) and any(w in lowered for w in trigger_words):
            return ".txt"
    return None


def _persist_document_artifact(
    coordinator: Any,
    content: str,
    trace_id: str,
    extension: str,
) -> str | None:
    """Write content to the artifact store and return the file path, or None on failure."""
    store = getattr(coordinator, "generated_artifact_store", None)
    if store is None or not content.strip():
        return None
    artifact_id = f"doc_summary_{trace_id}"
    try:
        path = store.persist("feature", artifact_id, content, extension=extension)
        return str(path)
    except Exception:
        return None


async def _invoke_document_model(coordinator: Any, *, task_type: str, prompt: str, system_prompt: str) -> str:
    router = getattr(coordinator, "model_router", None)
    if router is None:
        return ""
    try:
        return await router.invoke(task_type=task_type, prompt=prompt, system_prompt=system_prompt)
    except Exception:
        return ""


async def _invoke_visual_model(
    coordinator: Any,
    *,
    task_type: str,
    prompt: str,
    system_prompt: str,
    image_path: Path,
    content_type: str,
) -> str:
    router = getattr(coordinator, "model_router", None)
    if router is None:
        return ""
    try:
        return await router.invoke_multimodal(
            task_type=task_type,
            system_prompt=system_prompt,
            user_content=[
                {"type": "text", "text": prompt},
                {"type": "image_url", "image_url": {"url": _image_to_data_url(image_path, content_type)}},
            ],
        )
    except Exception:
        return ""


def _run_in_existing_loop(coro: Any) -> Any:
    import asyncio
    import threading

    result: dict[str, Any] = {}
    error: dict[str, Exception] = {}

    def _runner() -> None:
        try:
            result["value"] = asyncio.run(coro)
        except Exception as exc:  # pragma: no cover
            error["value"] = exc

    thread = threading.Thread(target=_runner, daemon=True)
    thread.start()
    thread.join()
    if "value" in error:
        raise error["value"]
    return result.get("value")


# ──────────────────────────────────────────────────────────────────────────────
# Compound pipeline: Step 1 – Summarize
# ──────────────────────────────────────────────────────────────────────────────

def handle_summarize_document_step(
    coordinator: Any,
    _step: dict[str, Any],
    task: TaskSchema,
    context: CoreContextSchema,
    _step_outputs: dict[str, Any],
) -> dict[str, Any]:
    """Step 1 of the compound pipeline: extract text and produce a concise summary."""
    targets = _resolve_analysis_targets(task.input_text, context)
    if not targets:
        return {
            "artifact_type": "document",
            "status": "no_document",
            "message": "当前 session 中没有可分析的文档，请先上传或发送文档。",
            "summary": "",
            "source_documents": [],
            "parsers": [],
        }

    extracted_chunks: list[str] = []
    source_files: list[str] = []
    parsers: list[str] = []
    for item in targets:
        path_value = str(item.get("received_path") or item.get("stored_path") or "").strip()
        file_name = str(item.get("file_name") or Path(path_value).name or "document")
        source_files.append(file_name)
        if not path_value or not Path(path_value).exists():
            continue
        if _is_image_target(item):
            continue
        content, parser_name = _extract_document_text(Path(path_value))
        parsers.append(parser_name)
        if content.strip():
            extracted_chunks.append(f"# {file_name}\n{content.strip()}")

    combined_text = "\n\n".join(extracted_chunks).strip()
    analysis_text = combined_text[:12000]
    model_output = _run_in_existing_loop(
        _invoke_document_model(
            coordinator,
            task_type="document_analysis",
            prompt=(
                "Summarize the following document content. "
                "Highlight the main points, important facts, and any action items.\n\n"
                f"{analysis_text}"
            ),
            system_prompt="You analyze user-provided documents. Return concise, structured plain text focused on the user's request.",
        )
    ) if analysis_text else ""

    summary = model_output.strip() or _fallback_summary(analysis_text, source_files)
    return {
        "artifact_type": "document",
        "status": "completed",
        "message": f"已完成文档总结，共处理 {len(source_files)} 个文件。",
        "summary": summary,
        "content": combined_text[:2000],
        "source_documents": source_files,
        "parsers": parsers,
        "document_count": len(source_files),
    }


# ──────────────────────────────────────────────────────────────────────────────
# Compound pipeline: Step 2 – Translate summary
# ──────────────────────────────────────────────────────────────────────────────

def handle_translate_summary_step(
    coordinator: Any,
    _step: dict[str, Any],
    task: TaskSchema,
    context: CoreContextSchema,
    step_outputs: dict[str, Any],
) -> dict[str, Any]:
    """Step 2 of the compound pipeline: translate the summary produced by Step 1."""
    prior = step_outputs.get("summarize_document") or {}
    text_to_translate = str(prior.get("summary") or "").strip()
    if not text_to_translate:
        return {
            "artifact_type": "document",
            "status": "no_content",
            "message": "无可翻译的摘要内容，请确认步骤 1 已成功完成。",
            "translation": "",
        }

    pending_request = _get_active_pending_request(context.session_state or {})
    target_language = str(
        task.constraints.get("target_language")
        or (pending_request or {}).get("target_language")
        or _detect_target_language(task.input_text, context)
    )

    model_output = _run_in_existing_loop(
        _invoke_document_model(
            coordinator,
            task_type="document_analysis",
            prompt=(
                f"Translate the following text into {target_language}. "
                "Preserve the structure and key information. "
                "Return only the translated text, no explanation.\n\n"
                f"{text_to_translate}"
            ),
            system_prompt=(
                "You translate user-provided text. Preserve the original meaning, "
                "keep structure when useful, and return plain text only."
            ),
        )
    )

    translation = model_output.strip() or text_to_translate
    return {
        "artifact_type": "document",
        "status": "completed",
        "message": f"已将摘要翻译为 {target_language}。",
        "translation": translation,
        "target_language": target_language,
        "source_summary": text_to_translate,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Compound pipeline: Step 3 – Save to file
# ──────────────────────────────────────────────────────────────────────────────

def handle_save_document_file_step(
    coordinator: Any,
    _step: dict[str, Any],
    task: TaskSchema,
    context: CoreContextSchema,
    step_outputs: dict[str, Any],
) -> dict[str, Any]:
    """Step 3 of the compound pipeline: persist the translated summary as a text file."""
    translate_out = step_outputs.get("translate_summary") or {}
    summarize_out = step_outputs.get("summarize_document") or {}

    file_content = (
        str(translate_out.get("translation") or "").strip()
        or str(summarize_out.get("summary") or "").strip()
    )
    if not file_content:
        return {
            "artifact_type": "document",
            "status": "no_content",
            "message": "没有可保存的内容，请确认前置步骤已成功完成。",
            "artifact_path": "",
        }

    extension = _detect_output_file_format(task.input_text) or ".txt"
    artifact_path = _persist_document_artifact(coordinator, file_content, context.trace_id, extension)
    session_store = getattr(coordinator, "session_store", None)
    pending_request = _get_active_pending_request(context.session_state or {})
    if session_store is not None and pending_request is not None:
        session_store.patch(context.session_id, {"pending_document_request": None})

    return {
        "artifact_type": "document",
        "status": "completed",
        "message": f"文件已生成：{Path(artifact_path).name if artifact_path else '(保存失败)'}",
        "artifact_path": artifact_path or "",
    }


def handle_analyze_document_step(
    coordinator: Any,
    _step: dict[str, Any],
    task: TaskSchema,
    context: CoreContextSchema,
    _step_outputs: dict[str, Any],
) -> dict[str, Any]:
    pending_request = _get_active_pending_request(context.session_state or {})
    wait_state = str(task.constraints.get("document_wait_state") or "").strip()
    session_store = getattr(coordinator, "session_store", None)

    if wait_state == "awaiting_document":
        now = _utc_now()
        timeout_seconds = _document_wait_timeout_seconds()
        expires_at = now + timedelta(seconds=timeout_seconds)
        if session_store is not None:
            session_store.patch(
                context.session_id,
                {
                    "pending_document_request": {
                        "request_text": str(task.constraints.get("document_request_text") or task.input_text or "").strip(),
                        "document_action": str(task.constraints.get("document_action") or "summarize"),
                        "target_language": str(task.constraints.get("target_language") or ""),
                        "requested_at": now.isoformat(),
                        "expires_at": expires_at.isoformat(),
                    }
                },
            )
        return {
            "artifact_type": "document",
            "status": "awaiting_document",
            "message": _build_standby_message(wait_state, context),
            "standby": True,
            "suppress_reply": _is_bridge_line_context(context),
            "wait_until": expires_at.isoformat(),
        }

    if wait_state == "awaiting_request":
        return {
            "artifact_type": "document",
            "status": "awaiting_request",
            "message": _build_standby_message(wait_state, context),
            "standby": True,
            "suppress_reply": _is_bridge_line_context(context),
        }

    targets = _resolve_analysis_targets(task.input_text, context)
    if not targets:
        return {
            "artifact_type": "document",
            "status": "no_document",
            "message": "当前 session 中没有可分析的文档，请先上传或发送文档。",
        }

    extracted_chunks: list[str] = []
    source_files: list[str] = []
    parsers: list[str] = []
    image_targets: list[tuple[Path, str, str]] = []
    for item in targets:
        path_value = str(item.get("received_path") or item.get("stored_path") or "").strip()
        file_name = str(item.get("file_name") or Path(path_value).name or "document")
        source_files.append(file_name)
        if not path_value or not Path(path_value).exists():
            continue
        if _is_image_target(item):
            image_targets.append((Path(path_value), str(item.get("content_type") or "image/png"), file_name))
        content, parser_name = _extract_document_text(Path(path_value))
        parsers.append(parser_name)
        if content.strip():
            extracted_chunks.append(f"# {file_name}\n{content.strip()}")

    combined_text = "\n\n".join(extracted_chunks).strip()
    operation = str(task.constraints.get("document_action") or (pending_request.get("document_action") if pending_request else "summarize"))
    if operation not in {"summarize", "translate", "analyze"}:
        operation = "summarize"
    target_language = str(task.constraints.get("target_language") or (pending_request or {}).get("target_language") or _detect_target_language(task.input_text, context))
    analysis_text = combined_text[:12000]
    use_visual_analysis = bool(image_targets) and operation == "analyze" and _looks_like_visual_analysis_request(task.input_text)

    if use_visual_analysis:
        system_prompt = (
            "You analyze image content for the user. Identify visible subjects, objects, animals, scenes, and other important details. "
            "Answer in concise plain text and do not fabricate details that are not visible."
        )
        prompt = f"Analyze this image and answer the user's request: {task.input_text.strip()}"
        image_path, content_type, _image_name = image_targets[-1]
        visual_output = _run_in_existing_loop(
            _invoke_visual_model(
                coordinator,
                task_type="image_understanding",
                prompt=prompt,
                system_prompt=system_prompt,
                image_path=image_path,
                content_type=content_type,
            )
        )
        summary = visual_output.strip() or _fallback_summary(analysis_text, source_files)
        if session_store is not None and pending_request is not None:
            session_store.patch(context.session_id, {"pending_document_request": None})
        return {
            "artifact_type": "document",
            "status": "completed",
            "message": f"已完成图片内容分析，共处理 {len(source_files)} 个文件。",
            "summary": summary,
            "translation": "",
            "content": combined_text[:2000] if combined_text else "",
            "source_documents": source_files,
            "parsers": parsers,
            "document_count": len(source_files),
            "requested_action": "analyze_visual",
            "target_language": "",
        }

    if operation == "translate":
        system_prompt = (
            "You translate user-provided documents. Preserve the original meaning, keep structure when useful, "
            "and return plain text only."
        )
        prompt = (
            f"Translate the following document content into {target_language}. "
            f"If the document is long, provide a faithful condensed translation.\n\n{analysis_text}"
        )
        model_output = _run_in_existing_loop(
            _invoke_document_model(coordinator, task_type="document_analysis", prompt=prompt, system_prompt=system_prompt)
        ) if analysis_text else ""
        translation = model_output.strip() or _fallback_summary(analysis_text, source_files)
        message = f"已完成文档翻译，共处理 {len(source_files)} 个文件。"
        summary = ""
    else:
        system_prompt = (
            "You analyze user-provided documents. Return concise, structured plain text focused on the user's request."
        )
        prompt = (
            "Summarize the following document content. Highlight the main points, important facts, and any action items.\n\n"
            f"{analysis_text}"
        )
        model_output = _run_in_existing_loop(
            _invoke_document_model(coordinator, task_type="document_analysis", prompt=prompt, system_prompt=system_prompt)
        ) if analysis_text else ""
        summary = model_output.strip() or _fallback_summary(analysis_text, source_files)
        translation = ""
        message = f"已完成文档总结，共处理 {len(source_files)} 个文件。"

    if session_store is not None and pending_request is not None:
        session_store.patch(context.session_id, {"pending_document_request": None})

    output_file_ext = _detect_output_file_format(task.input_text)
    artifact_path: str | None = None
    if output_file_ext:
        file_content = summary or translation
        artifact_path = _persist_document_artifact(coordinator, file_content, context.trace_id, output_file_ext)

    return {
        "artifact_type": "document",
        "status": "completed",
        "message": message,
        "summary": summary,
        "translation": translation,
        "content": combined_text[:2000] if combined_text else "",
        "source_documents": source_files,
        "parsers": parsers,
        "document_count": len(source_files),
        "requested_action": operation,
        "target_language": target_language if operation == "translate" else "",
        "artifact_path": artifact_path or "",
    }


class DocumentIntentPlugin:
    priority = 110

    def match(self, text: str, context: CoreContextSchema) -> bool:
        current_attachments = list((context.metadata or {}).get("analysis_attachments") or [])
        session_attachments = list((context.session_state or {}).get("analysis_attachments") or [])
        has_context = bool(current_attachments or session_attachments or _get_active_pending_request(context.session_state or {}))
        if current_attachments:
            return True
        return _looks_like_explicit_document_request(text, has_context=has_context)

    def run(self, text: str, context: CoreContextSchema) -> dict[str, Any]:
        current_attachments = list((context.metadata or {}).get("analysis_attachments") or [])
        session_attachments = list((context.session_state or {}).get("analysis_attachments") or [])
        pending_request = _get_active_pending_request(context.session_state or {})
        has_context = bool(current_attachments or session_attachments)
        explicit_request = _looks_like_explicit_document_request(text, has_context=has_context)
        operation = _detect_document_action(text if explicit_request else str((pending_request or {}).get("request_text") or text))
        constraints: dict[str, Any] = {
            "need_agent": False,
            "document_action": operation,
        }
        if explicit_request and operation == "translate":
            constraints["target_language"] = _detect_target_language(text, context)
        elif pending_request and str(pending_request.get("document_action") or "") == "translate":
            constraints["target_language"] = str(pending_request.get("target_language") or _detect_target_language(text, context))

        if explicit_request and not (current_attachments or session_attachments):
            constraints["document_wait_state"] = "awaiting_document"
            constraints["document_request_text"] = text.strip()
            constraints["suppress_reply"] = _is_bridge_line_context(context)
        elif current_attachments and not explicit_request and pending_request:
            constraints["consume_pending_request"] = True
        elif current_attachments and not explicit_request:
            constraints["document_wait_state"] = "awaiting_request"
            constraints["suppress_reply"] = _is_bridge_line_context(context)

        analysis = {
            "document_context": True,
            "document_count": len((context.session_state or {}).get("documents") or []),
            "analysis_attachment_count": len(session_attachments or current_attachments),
            "document_action": operation,
            "wait_state": str(constraints.get("document_wait_state") or ""),
        }
        outputs = ["text", "summary"] if operation != "translate" else ["text", "translation"]
        return {
            "intent": "file_upload_task",
            "domain": "multimodal_ops",
            "output_requirements": outputs,
            "constraints": constraints,
            "analysis": analysis,
        }


class DocumentTaskDecomposerPlugin:
    priority = 110

    def match(self, task: TaskSchema) -> bool:
        return task.intent == "file_upload_task" and "document_action" in task.constraints

    def run(self, task: TaskSchema) -> list[SubTask]:
        action = str(task.constraints.get("document_action") or "summarize")
        if action in _COMPOUND_ACTIONS:
            steps_meta = [
                ("summarize_document",   "Step 1: Summarize the uploaded document."),
                ("translate_summary",    "Step 2: Translate the summary into the requested language."),
                ("save_document_file",   "Step 3: Save the translated summary as a text file."),
            ]
            return [
                SubTask(subtask_id=generate_id("subtask"), name=name, goal=goal)
                for name, goal in steps_meta
            ]
        return [
            SubTask(
                subtask_id=generate_id("subtask"),
                name="analyze_document",
                goal="Analyze the latest session document or the referenced uploaded documents.",
            )
        ]


class DocumentWorkflowPlannerPlugin:
    priority = 110

    def match(self, task: TaskSchema, _subtasks: list[SubTask]) -> bool:
        return task.intent == "file_upload_task" and "document_action" in task.constraints

    def run(self, task: TaskSchema, _subtasks: list[SubTask]) -> WorkflowSchema:
        action = str(task.constraints.get("document_action") or "summarize")

        if action in _COMPOUND_ACTIONS:
            step_summarize = WorkflowStepSchema(
                step_id=generate_id("step"),
                name="summarize_document",
                task_type=task.intent,
                executor_type="tool",
                inputs=["input_text", "session_state", "attachments"],
                outputs=["summary", "content", "source_documents", "parsers"],
                depends_on=[],
                retry=0,
                metadata={
                    "goal": "Extract text from the uploaded document and produce a concise summary.",
                    "display_label": "📄 文档总结",
                },
            )
            step_translate = WorkflowStepSchema(
                step_id=generate_id("step"),
                name="translate_summary",
                task_type=task.intent,
                executor_type="tool",
                inputs=["summarize_document.summary"],
                outputs=["translation", "target_language"],
                depends_on=["summarize_document"],
                retry=0,
                metadata={
                    "goal": "Translate the summary produced in step 1 into the target language.",
                    "display_label": "🌐 翻译摘要",
                },
            )
            step_save = WorkflowStepSchema(
                step_id=generate_id("step"),
                name="save_document_file",
                task_type=task.intent,
                executor_type="tool",
                inputs=["translate_summary.translation"],
                outputs=["artifact_path"],
                depends_on=["translate_summary"],
                retry=0,
                metadata={
                    "goal": "Save the translated text as a downloadable text file.",
                    "display_label": "💾 生成文件",
                },
            )
            steps = [step_summarize, step_translate, step_save]
            if action == "summarize_and_translate":
                steps = [step_summarize, step_translate]
        else:
            steps = [
                WorkflowStepSchema(
                    step_id=generate_id("step"),
                    name="analyze_document",
                    task_type=task.intent,
                    executor_type="tool",
                    inputs=["input_text", "session_state", "attachments"],
                    outputs=["message", "summary", "translation", "content", "source_documents"],
                    depends_on=[],
                    retry=0,
                    metadata={
                        "goal": "Summarize, analyze, or translate the uploaded document within the current session.",
                        "selection_basis": "document_runtime_plugin",
                        "display_label": "📄 文档处理",
                    },
                )
            ]

        return WorkflowSchema(
            workflow_id=generate_id("workflow"),
            task_id=task.task_id,
            mode="normal",
            steps=steps,
            composition={
                "plugin": "document_runtime_plugin",
                "intent": task.intent,
                "document_action": action,
            },
        )


class DocumentExecutionHandlerPlugin:
    def build_manifest(self, _coordinator: Any) -> ExecutionHandlerPluginManifest:
        return ExecutionHandlerPluginManifest(
            name="document_runtime_plugin",
            version="1.0",
            steps=[
                ExecutionHandlerPluginStepSpec(
                    executor_type="tool",
                    step_name="analyze_document",
                    handler=lambda step, task, context, step_outputs: handle_analyze_document_step(
                        _coordinator, step, task, context, step_outputs
                    ),
                    description="Summarize or translate session-bound uploaded documents (single-action).",
                ),
                ExecutionHandlerPluginStepSpec(
                    executor_type="tool",
                    step_name="summarize_document",
                    handler=lambda step, task, context, step_outputs: handle_summarize_document_step(
                        _coordinator, step, task, context, step_outputs
                    ),
                    description="Step 1 of compound pipeline: extract text and summarize the document.",
                ),
                ExecutionHandlerPluginStepSpec(
                    executor_type="tool",
                    step_name="translate_summary",
                    handler=lambda step, task, context, step_outputs: handle_translate_summary_step(
                        _coordinator, step, task, context, step_outputs
                    ),
                    description="Step 2 of compound pipeline: translate the summary.",
                ),
                ExecutionHandlerPluginStepSpec(
                    executor_type="tool",
                    step_name="save_document_file",
                    handler=lambda step, task, context, step_outputs: handle_save_document_file_step(
                        _coordinator, step, task, context, step_outputs
                    ),
                    description="Step 3 of compound pipeline: save the result as a downloadable file.",
                ),
            ],
        )


def document_intent_plugin() -> DocumentIntentPlugin:
    return DocumentIntentPlugin()


def document_task_decomposer_plugin() -> DocumentTaskDecomposerPlugin:
    return DocumentTaskDecomposerPlugin()


def document_workflow_planner_plugin() -> DocumentWorkflowPlannerPlugin:
    return DocumentWorkflowPlannerPlugin()


def document_execution_handler_plugin() -> DocumentExecutionHandlerPlugin:
    return DocumentExecutionHandlerPlugin()